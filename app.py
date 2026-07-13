# =============================================================================
# reports_app/app.py  –  Thinkster Math AI Reports Application
# =============================================================================
# Run with:  python app.py   (inside reports_app/)
# =============================================================================

import os
import io
import re
from datetime import datetime

from dotenv import load_dotenv

# Load .env from the same directory as this file
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

from flask import Flask, render_template, jsonify, request, send_file, session, redirect, url_for
from pymongo import MongoClient
from werkzeug.security import generate_password_hash, check_password_hash


# ---------------------------------------------------------------------------
# Config from .env
# ---------------------------------------------------------------------------
MONGO_URI   = os.getenv("MONGO_URI", "")
MONGO_DB    = os.getenv("MONGO_DB", "Thinkster_testing")
SECRET_KEY  = os.getenv("SECRET_KEY", "dev_secret")

if not MONGO_URI:
    raise RuntimeError("MONGO_URI is not set. Check your .env file.")

# ---------------------------------------------------------------------------
# Flask setup
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = SECRET_KEY

# ---------------------------------------------------------------------------
# MongoDB setup
# ---------------------------------------------------------------------------
mongo_client    = MongoClient(MONGO_URI)
db              = mongo_client[MONGO_DB]
report_col      = db["Worksheet_Report"]      # AI analysis results
ws_answers_col  = db["WS_answers"]            # topic names + answer keys
answering_col   = db["Answering_Report"]      # answering run reports
users_col       = db["users"]                 # user credentials


 

def _topic_map() -> dict:
    """Return {worksheet_id: topic_name} from WS_answers collection."""
    mapping = {}
    for doc in ws_answers_col.find({}, {"worksheetID": 1, "topicName": 1}):
        ws_id  = doc.get("worksheetID", "")
        topic  = doc.get("topicName", "Unknown Topic") or "Unknown Topic"
        if ws_id:
            mapping[ws_id] = topic
    return mapping


def _get_all_reports():
    """
    Aggregate Worksheet_Report by worksheet_id.
    Returns a list of dicts:
      { worksheet_id, topic_name, total_questions,
        issue_count, passed_count, has_issue, questions: [...] }
    """
    topic_map = _topic_map()
    pipeline = [
        {"$group": {
            "_id": "$worksheet_id",
            "total_questions": {"$sum": 1},
            "issue_count":     {"$sum": {"$cond": [{"$and": [{"$eq": ["$status", "Issue"]}, {"$ne": ["$solved", True]}]}, 1, 0]}},
            "passed_count":    {"$sum": {"$cond": [{"$or": [{"$eq": ["$status", "Passed"]}, {"$eq": ["$solved", True]}]}, 1, 0]}},
            "latest_time":     {"$max": "$analysis_time"},
        }},
        {"$sort": {"_id": 1}}
    ]
    rows = list(report_col.aggregate(pipeline))
    result = []
    # Fetch all worksheet IDs from answering_col to mark them
    auto_ans_ws_ids = set(answering_col.distinct("worksheet_id"))
    for row in rows:
        ws_id = row["_id"] or ""
        result.append({
            "worksheet_id":   ws_id,
            "topic_name":     topic_map.get(ws_id, "Unknown Topic"),
            "total_questions":row["total_questions"],
            "issue_count":    row["issue_count"],
            "passed_count":   row["passed_count"],
            "has_issue":      row["issue_count"] > 0,
            "latest_time":    row.get("latest_time", ""),
            "auto_answered":  ws_id in auto_ans_ws_ids,
        })
    return result


def _get_worksheet_detail(ws_id: str):
    """Return full question-level data for one worksheet, including correct answer from WS_answers database."""
    docs = list(report_col.find({"worksheet_id": ws_id}).sort("question_number", 1))
    topic_map = _topic_map()
    
    # Query WS_answers to fetch correct answers from the DB
    ws_answers_doc = ws_answers_col.find_one({"worksheetID": ws_id})
    db_answers_map = {}
    if ws_answers_doc:
        for k, v in ws_answers_doc.items():
            if k.startswith("q") and k[1:].isdigit():
                db_answers_map[int(k[1:])] = v
            
    # Check if worksheet was completed using automation and get student answers
    answering_doc = answering_col.find_one({"worksheet_id": ws_id})
    is_auto = answering_doc is not None
    answering_map = {}
    if answering_doc and "questions" in answering_doc:
        for q in answering_doc["questions"]:
            q_num = q.get("question_number")
            if q_num is not None:
                answering_map[int(q_num)] = {
                    "website_correct_answer": q.get("website_correct_answer", ""),
                    "submitted_answer": q.get("submitted_answer", "")
                }

    questions = []
    for d in docs:
        q_num = d.get("question_number", 0)
        ans_info = answering_map.get(q_num, {})
        questions.append({
            "question_number": q_num,
            "image_name":      d.get("image_name", ""),
            "ai_response":     d.get("ai_response", ""),
            "status":          d.get("status", "Unknown"),
            "solved":          d.get("solved", False),
            "analysis_time":   d.get("analysis_time", ""),
            "db_correct_answer": db_answers_map.get(q_num, ""),
            "website_correct_answer": ans_info.get("website_correct_answer", ""),
            "submitted_answer": ans_info.get("submitted_answer", ""),
        })
    return {
        "worksheet_id": ws_id,
        "topic_name":   topic_map.get(ws_id, "Unknown Topic"),
        "auto_answered": is_auto,
        "questions":    questions,
    }


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.before_request
def require_login():
    """Ensure user is logged in before accessing protected routes."""
    # Allow login and static assets to bypass check
    if request.endpoint in ["login", "static"]:
        return None
        
    if not session.get("logged_in"):
        if request.path.startswith("/api/"):
            return jsonify({"error": "Unauthorized"}), 401
        return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("logged_in"):
        return redirect(url_for("index"))
        
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        
        user = users_col.find_one({"username": username})
        if user and check_password_hash(user["password"], password):
            session["logged_in"] = True
            session["username"] = username
            return redirect(url_for("index"))
        else:
            error = "Invalid username or password"
            
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/reports")
def api_reports():
    """Return all aggregated report rows."""
    reports = _get_all_reports()
    return jsonify(reports)


@app.route("/api/topics")
def api_topics():
    """Return sorted unique topic names."""
    reports = _get_all_reports()
    topics = sorted({r["topic_name"] for r in reports if r["topic_name"] != "Unknown Topic"})
    return jsonify(topics)


@app.route("/api/topic-summary")
def api_topic_summary():
    """Return summary stats for a given topic."""
    topic = request.args.get("topic", "").strip()
    reports = _get_all_reports()
    if topic:
        rows = [r for r in reports if r["topic_name"] == topic]
    else:
        rows = reports
    total        = len(rows)
    with_issues  = sum(1 for r in rows if r["has_issue"])
    without_issues = total - with_issues
    return jsonify({
        "total":           total,
        "with_issues":     with_issues,
        "without_issues":  without_issues,
    })


@app.route("/api/worksheet/<ws_id>")
def api_worksheet_detail(ws_id):
    """Return detailed question-level data for one worksheet."""
    data = _get_worksheet_detail(ws_id)
    return jsonify(data)


@app.route("/api/worksheet/mark-solved", methods=["POST"])
def api_mark_solved():
    """Mark a specific question in a worksheet as solved."""
    body = request.get_json(force=True) or {}
    ws_id = body.get("worksheet_id")
    q_num = body.get("question_number")
    if not ws_id or q_num is None:
        return jsonify({"error": "Missing worksheet_id or question_number"}), 400
    
    # Update status/solved status in Worksheet_Report
    res = report_col.update_one(
        {"worksheet_id": ws_id, "question_number": int(q_num)},
        {"$set": {"solved": True}}
    )
    if res.modified_count == 0:
        doc = report_col.find_one({"worksheet_id": ws_id, "question_number": int(q_num)})
        if not doc:
            return jsonify({"error": "Question not found"}), 404
            
    return jsonify({"success": True})


def generate_pdf_table_html(title, worksheet_rows, max_q, stats_data=None):
    header_cols = "".join(f"<th>Q{i}</th>" for i in range(1, max_q + 1))
    
    rows_html = ""
    for ws_id, qs in sorted(worksheet_rows.items()):
        row_cells = f"<td style='font-weight:bold; color: #111827;'>{ws_id}</td>"
        for i in range(1, max_q + 1):
            cell_val = qs.get(f"Q{i}", "")
            if cell_val:
                is_issue = "Issue:" in cell_val or "status\":\"Issue\"" in cell_val or "status\": \"Issue\"" in cell_val
                # Check fuzzy matches if the text starts with Issue or contains Issue
                if not is_issue:
                    is_issue = cell_val.strip().lower().startswith("issue")
                badge_class = 'badge-danger' if is_issue else 'badge-success'
                badge_text = 'Issue' if is_issue else 'Passed'
                badge = f"<span class='badge {badge_class}'>{badge_text}</span>"
                row_cells += f"<td>{badge}<div class='content'>{cell_val}</div></td>"
            else:
                row_cells += "<td><span style='color: #6b7280;'>N/A</span></td>"
        rows_html += f"<tr>{row_cells}</tr>"

    stats_html = ""
    if stats_data:
        stats_html = f"""
        <div class="summary">
            <div class="card"><div class="lbl">Total Worksheets</div><div class="val">{stats_data.get('total_ws', 0)}</div></div>
            <div class="card"><div class="lbl">Total Questions</div><div class="val">{stats_data.get('total_q', 0)}</div></div>
            <div class="card"><div class="lbl">Issues Found</div><div class="val">{stats_data.get('issues_count', 0)}</div></div>
            <div class="card"><div class="lbl">Passed Questions</div><div class="val">{stats_data.get('passed_count', 0)}</div></div>
            <div class="card"><div class="lbl">Export Date</div><div class="val">{datetime.now().strftime("%Y-%m-%d %H:%M")}</div></div>
        </div>
        """

    pdf_print_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{title}</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    <script>
        window.MathJax = {{
            tex: {{
                inlineMath: [['$', '$'], ['\\\\(', '\\\\)']],
                displayMath: [['$$', '$$'], ['\\\\[', '\\\\]']]
            }}
        }};
    </script>
    <script src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-mml-chtml.js" async></script>
    <style>
        body {{
            font-family: 'Inter', sans-serif;
            margin: 30px;
            background: #F9FAFB;
            color: #111827;
        }}
        h1 {{
            font-size: 24px;
            font-weight: 800;
            color: #111827;
            border-bottom: 2px solid #E5E7EB;
            padding-bottom: 10px;
            margin-bottom: 20px;
        }}
        .summary {{ display: flex; gap: 15px; margin-bottom: 30px; flex-wrap: wrap; }}
        .card {{
            background: #FFFFFF;
            border: 1px solid #E5E7EB;
            padding: 15px;
            border-radius: 8px;
            min-width: 140px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }}
        .card .lbl {{ font-size: 11px; color: #6B7280; text-transform: uppercase; letter-spacing: 0.5px; }}
        .card .val {{ font-size: 16px; font-weight: bold; margin-top: 5px; color: #111827; }}
        table {{
            width: 100%;
            border-collapse: collapse;
            background: #FFFFFF;
            border: 1px solid #E5E7EB;
            border-radius: 8px;
            overflow: hidden;
            font-size: 13px;
        }}
        th, td {{
            padding: 12px 14px;
            border: 1px solid #E5E7EB;
            vertical-align: top;
            text-align: left;
            color: #374151;
        }}
        th {{
            background: #F3F4F6;
            font-weight: 600;
            color: #111827;
        }}
        .badge {{
            display: inline-block;
            padding: 2px 6px;
            border-radius: 4px;
            font-size: 10px;
            font-weight: bold;
            margin-bottom: 6px;
        }}
        .badge-success {{ background: #D1FAE5; color: #065F46; border: 1px solid #A7F3D0; }}
        .badge-danger {{ background: #FEE2E2; color: #991B1B; border: 1px solid #FECACA; }}
        .content {{ font-size: 12px; white-space: pre-line; color: #4B5563; }}
        
        .no-print-btn {{
            font-family: inherit;
            font-weight: 600;
            font-size: 14px;
            padding: 8px 16px;
            border-radius: 8px;
            border: 1px solid #E5E7EB;
            background: #1F2937;
            color: white;
            cursor: pointer;
            box-shadow: 0 1px 2px rgba(0,0,0,0.05);
            transition: all 0.2s;
        }}
        .no-print-btn:hover {{
            background: #111827;
        }}
        @media print {{
            body {{
                background: white !important;
                color: black !important;
                margin: 0;
            }}
            h1 {{
                color: black !important;
                border-bottom: 2px solid #ddd;
            }}
            .card {{
                background: white !important;
                border: 1px solid #ddd !important;
                box-shadow: none !important;
            }}
            .card .val {{ color: black !important; }}
            table {{
                background: white !important;
                border: 1px solid #ddd !important;
                color: black !important;
            }}
            th, td {{
                border: 1px solid #ddd !important;
                color: black !important;
            }}
            th {{ background: #f3f4f6 !important; }}
            .badge-success {{ background: #d1fae5 !important; color: #065f46 !important; border: 1px solid #a7f3d0 !important; }}
            .badge-danger {{ background: #fee2e2 !important; color: #991b1b !important; border: 1px solid #fecaca !important; }}
            .content {{ color: #374151 !important; }}
            .no-print {{ display: none; }}
            table {{ page-break-inside: auto; }}
            tr {{ page-break-inside: avoid; page-break-after: auto; }}
            thead {{ display: table-header-group; }}
        }}
    </style>
</head>
<body>
    <div style="display: flex; justify-content: space-between; align-items: center; border-bottom: 2px solid #E5E7EB; padding-bottom: 10px; margin-bottom: 20px;">
        <h1 style="border: none; margin: 0;">{title}</h1>
        <button class="no-print no-print-btn" onclick="window.print()">
            Print Report / Save to PDF
        </button>
    </div>
    
    {stats_html}
    
    <table>
        <thead>
            <tr>
                <th>Worksheet ID</th>
                {header_cols}
            </tr>
        </thead>
        <tbody>
            {rows_html}
        </tbody>
    </table>
    <script>
        window.onload = function() {{
            setTimeout(function() {{
                window.print();
            }}, 1200);
        }};
    </script>
</body>
</html>
"""
    return pdf_print_content


@app.route("/api/export/pdf", methods=["POST"])
def export_pdf():
    """
    Export selected worksheet reports in the matrix format used by the Automation App.
    Accepts JSON body: { "worksheet_ids": [...], "topic": "..." }
    """
    body        = request.get_json(force=True) or {}
    ws_ids      = body.get("worksheet_ids", [])
    topic_filter= body.get("topic", "")

    reports = _get_all_reports()
    if ws_ids:
        rows = [r for r in reports if r["worksheet_id"] in ws_ids]
    elif topic_filter:
        rows = [r for r in reports if r["topic_name"] == topic_filter]
    else:
        rows = reports

    matching_ids = [r["worksheet_id"] for r in rows]

    # Query all question reviews from Worksheet_Report
    docs = list(report_col.find({"worksheet_id": {"$in": matching_ids}}))
    
    # Group answers by worksheet ID and find maximum question number
    ws_answers = {}
    max_q = 0
    total_q = 0
    issues_count = 0
    passed_count = 0
    
    for doc in docs:
        ws_id = doc["worksheet_id"]
        q_num = doc.get("question_number", 0)
        if q_num > max_q:
            max_q = q_num
        
        if ws_id not in ws_answers:
            ws_answers[ws_id] = {}
        
        ai_res = doc.get("ai_response", "")
        ws_answers[ws_id][f"Q{q_num}"] = ai_res
        total_q += 1
        if doc.get("status", "Passed") == "Issue":
            issues_count += 1
        else:
            passed_count += 1

    stats = {
        "total_ws": len(ws_answers),
        "total_q": total_q,
        "issues_count": issues_count,
        "passed_count": passed_count
    }

    html_content = generate_pdf_table_html("Worksheet AI Analysis Report", ws_answers, max_q, stats)
    buf = io.BytesIO(html_content.encode("utf-8"))
    buf.seek(0)
    filename = f"thinkster_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
    return send_file(buf, mimetype="text/html",
                     as_attachment=True, download_name=filename)


@app.route("/api/export/excel", methods=["POST"])
def export_excel():
    """Export selected worksheet reports to Excel in grid matrix format."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    body        = request.get_json(force=True) or {}
    ws_ids      = body.get("worksheet_ids", [])
    topic_filter= body.get("topic", "")

    reports = _get_all_reports()
    if ws_ids:
        rows = [r for r in reports if r["worksheet_id"] in ws_ids]
    elif topic_filter:
        rows = [r for r in reports if r["topic_name"] == topic_filter]
    else:
        rows = reports

    matching_ids = [r["worksheet_id"] for r in rows]

    # Query Worksheet_Report
    docs = list(report_col.find({"worksheet_id": {"$in": matching_ids}}))
    
    ws_answers = {}
    max_q = 0
    total_q = 0
    issues_count = 0
    passed_count = 0
    
    for doc in docs:
        ws_id = doc["worksheet_id"]
        q_num = doc.get("question_number", 0)
        if q_num > max_q:
            max_q = q_num
        
        if ws_id not in ws_answers:
            ws_answers[ws_id] = {}
        
        ai_res = doc.get("ai_response", "")
        ws_answers[ws_id][f"Q{q_num}"] = ai_res
        total_q += 1
        if doc.get("status", "Passed") == "Issue":
            issues_count += 1
        else:
            passed_count += 1

    stats = {
        "total_ws": len(ws_answers),
        "total_q": total_q,
        "issues_count": issues_count,
        "passed_count": passed_count
    }

    # Load correct answers from WS_answers
    ws_db_answers = {}
    ws_answers_docs = list(ws_answers_col.find({"worksheetID": {"$in": matching_ids}}))
    for doc in ws_answers_docs:
        ws_id = doc["worksheetID"]
        ws_db_answers[ws_id] = {}
        for k, v in doc.items():
            if k.startswith("q") and k[1:].isdigit():
                ws_db_answers[ws_id][f"Q{k[1:]}"] = v

    wb = Workbook()

    # Shared style tokens
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # 1. AI Analysis Report Sheet
    ws1 = wb.active
    ws1.title = "AI Analysis Report"
    ws1.row_dimensions[1].height = 28
    ws1.column_dimensions["A"].width = 25
    
    headers1 = ["Worksheet ID"] + [f"Q{i}" for i in range(1, max_q + 1)]
    for col_idx, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
        if col_idx > 1:
            ws1.column_dimensions[get_column_letter(col_idx)].width = 45

    for r_idx, ws_id in enumerate(sorted(ws_answers.keys()), 2):
        ws1.row_dimensions[r_idx].height = 70
        cell_id = ws1.cell(row=r_idx, column=1, value=ws_id)
        cell_id.font      = Font(bold=True)
        cell_id.alignment = Alignment(horizontal="center", vertical="center")
        cell_id.border    = thin_border
        for q_idx in range(1, max_q + 1):
            val = ws_answers[ws_id].get(f"Q{q_idx}", "")
            cell = ws1.cell(row=r_idx, column=q_idx + 1, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border    = thin_border
            is_issue = "Issue:" in val or val.strip().lower().startswith("issue")
            if is_issue:
                cell.font = Font(color="991B1B", bold=True)
            elif val:
                cell.font = Font(color="065F46")

    # 2. Correct Answers (DB) Sheet
    ws2 = wb.create_sheet("Correct Answers (DB)")
    ws2.row_dimensions[1].height = 28
    ws2.column_dimensions["A"].width = 25
    for col_idx, header in enumerate(headers1, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
        if col_idx > 1:
            ws2.column_dimensions[get_column_letter(col_idx)].width = 25

    for r_idx, ws_id in enumerate(sorted(ws_answers.keys()), 2):
        ws2.row_dimensions[r_idx].height = 24
        cell_id = ws2.cell(row=r_idx, column=1, value=ws_id)
        cell_id.font      = Font(bold=True)
        cell_id.alignment = Alignment(horizontal="center", vertical="center")
        cell_id.border    = thin_border
        for q_idx in range(1, max_q + 1):
            val = ws_db_answers.get(ws_id, {}).get(f"Q{q_idx}", "")
            cell = ws2.cell(row=r_idx, column=q_idx + 1, value=str(val))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin_border

    # 3. Summary Sheet
    ws4 = wb.create_sheet("Summary")
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 18
    summary_data = [
        ("Total Worksheets",        len(ws_answers)),
        ("Total Questions Analysed",stats["total_q"]),
        ("Total Issues Found",      stats["issues_count"]),
        ("Total Passed",            stats["passed_count"]),
        ("Export Date",             datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for row_idx, (label, val) in enumerate(summary_data, 1):
        ws4.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws4.cell(row=row_idx, column=2, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"thinkster_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("=" * 60)
    print("  Thinkster Math – Reports Application")
    print("  http://127.0.0.1:5050")
    print("=" * 60)
    app.run(debug=True, host="0.0.0.0", port=5050)
